from __future__ import annotations

import asyncio
import datetime
import logging
import time
from collections import defaultdict
from decimal import Decimal
from functools import lru_cache, partial
from io import BytesIO
from typing import List, Dict, Annotated, Self, Generator, Optional, Tuple

import pycountry
from bson import ObjectId
from fastapi import APIRouter, status, Body, Depends, Query, HTTPException
from fastapi.responses import RedirectResponse, StreamingResponse, Response
from ics import Calendar, Event
from ics.grammar.parse import ContentLine
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pycountry.db import Country
from pydantic import BaseModel, Field, computed_field, EmailStr, PrivateAttr, field_serializer
from pydantic.functional_validators import field_validator, model_validator
from starlette.concurrency import run_in_threadpool

from ..dependencies import get_current_active_user_check_tenant, get_tenant, mongo_to_pydantic, tenant_var
from ..model import Team, TeamMember, get_unique_countries, DayType, User, Tenant, DayEntry, DayAudit
from ..notification_types import (
    ensure_valid_notification_types,
    list_notification_type_ids,
    list_notification_types,
)
from ..routers.daytypes import DayTypeReadDTO, get_all_day_types
from ..routers.users import UserWithoutTenantsDTO
from ..utils import get_country_holidays
from ..utils import get_today

log = logging.getLogger(__name__)
router = APIRouter(prefix="/teams", tags=["Teams"])


def validate_country_name(country_name):
    countries = pycountry.countries.search_fuzzy(country_name)
    if countries:
        country = countries[0]
        if isinstance(country, Country):
            return country.name
    return None


class TeamMemberWriteDTO(BaseModel):
    name: str
    country: str
    email: EmailStr | None = None
    phone: str | None = None
    available_day_types: List[DayTypeReadDTO] = []
    birthday: str | None = Field(None, pattern=r"^(0[1-9]|1[0-2])-(0[1-9]|[12][0-9]|3[01])$")
    employee_start_date: datetime.date | None = None
    yearly_vacation_days: Decimal | None = None

    @field_validator("country")
    @classmethod
    def validate_country(cls, value: str) -> str:
        country_name = validate_country_name(value)
        if country_name:
            return country_name
        raise ValueError("Invalid country name")

    @field_validator('email', mode='before')
    @classmethod
    def empty_email_to_none(cls, v):
        return None if v == "" else v

    @field_validator('birthday', mode='before')
    @classmethod
    def empty_birthday_to_none(cls, v):
        return None if v == "" else v


class DayEntryDTO(BaseModel):
    day_types: List[DayTypeReadDTO] = []
    comment: str = ''


class TeamMemberReadDTO(TeamMemberWriteDTO):
    uid: str
    days: Dict[str, DayEntryDTO] = Field(default_factory=dict)
    is_deleted: bool = False
    deleted_at: datetime.datetime | None = None
    deleted_by: UserWithoutTenantsDTO | None = None
    _vacation_split_cache: Optional[Tuple[Dict[int, int], Dict[int, int]]] = PrivateAttr(default=None)

    def _split_vacation_days(self) -> tuple[Dict[int, int], Dict[int, int]]:
        """Return two dicts: used days and planned days by year."""
        if self._vacation_split_cache is not None:
            return self._vacation_split_cache

        used = defaultdict(int)
        planned = defaultdict(int)
        vacation_day_type = mongo_to_pydantic(
            DayType.objects(tenant=tenant_var.get(), name="Vacation").first(),
            DayTypeReadDTO,
        )
        today = get_today()
        for date_str, day_entry in self.days.items():
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            day_types = day_entry.day_types
            if vacation_day_type in day_types:
                if date <= today:
                    used[date.year] += 1
                else:
                    planned[date.year] += 1

        self._vacation_split_cache = (dict(used), dict(planned))
        return self._vacation_split_cache

    @computed_field
    @property
    def vacation_used_days_by_year(self) -> Dict[int, int]:
        used, _ = self._split_vacation_days()
        return used

    @computed_field
    @property
    def vacation_planned_days_by_year(self) -> Dict[int, int]:
        _, planned = self._split_vacation_days()
        return planned

    @computed_field
    @property
    def vacation_available_days(self) -> int | None:
        if self.yearly_vacation_days is None or self.employee_start_date is None:
            return None

        yearly = Decimal(str(self.yearly_vacation_days))
        start = self.employee_start_date
        today = get_today()

        total_budget = Decimal("0")
        for year in range(start.year, today.year + 1):
            if year == start.year:
                days_in_year = (datetime.date(year, 12, 31) - datetime.date(year, 1, 1)).days + 1
                days_employed = (datetime.date(year, 12, 31) - start).days + 1
                portion = (Decimal(days_employed) / Decimal(days_in_year)) * yearly
                total_budget += portion
            else:
                total_budget += yearly

        used_total = sum(
            count for year, count in self.vacation_used_days_by_year.items()
            if year <= today.year
        )
        planned_total = sum(
            count for year, count in self.vacation_planned_days_by_year.items()
            if year <= today.year
        )

        available = int(total_budget - used_total - planned_total)
        return max(0, available)

    @model_validator(mode='after')
    def include_birthday(self) -> Self:
        def add_birthday(year):
            birthday_date = f"{year}-{self.birthday}"
            if birthday_date not in self.days:
                self.days[birthday_date] = DayEntryDTO()
            self.days[birthday_date].day_types.append(
                DayTypeReadDTO.from_mongo_reference_field(DayType.get_birthday_day_type_id(tenant_var.get())))

        if self.birthday:
            current_year = datetime.datetime.now().year
            add_birthday(current_year)
            add_birthday(current_year + 1)

        return self

    @field_validator('days', mode="before")
    @classmethod
    def convert_days(cls, v):
        if v and isinstance(v, dict):
            converted_days = v
            for date_str, day_entry in v.items():
                day_type_ids = day_entry["day_types"]
                converted_day_types = []
                for day_type_id in day_type_ids:
                    if isinstance(day_type_id, ObjectId):
                        converted_day_types.append(DayTypeReadDTO.from_mongo_reference_field(day_type_id))
                converted_days[date_str]["day_types"] = converted_day_types
            return converted_days
        return v

    @field_validator('deleted_by', mode="before")
    @classmethod
    def convert_deleted_by(cls, value):
        if isinstance(value, ObjectId):
            return UserWithoutTenantsDTO.from_mongo_reference_field(value)
        if isinstance(value, User):
            return mongo_to_pydantic(value, UserWithoutTenantsDTO)
        return value

    @computed_field
    @property
    def country_flag(self) -> str:
        country = pycountry.countries.get(name=self.country)
        if country:
            return country.flag
        raise ValueError(f"Invalid country name: {self.country}")


class TeamWriteDTO(BaseModel):
    name: str
    available_day_types: List[DayTypeReadDTO] = []


class TeamReadDTO(TeamWriteDTO):
    id: str = Field(None, alias='_id')
    team_members: List[TeamMemberReadDTO]
    subscribers: List[UserWithoutTenantsDTO] = []
    is_deleted: bool = False
    deleted_at: datetime.datetime | None = None
    deleted_by: UserWithoutTenantsDTO | None = None

    @field_validator('team_members')
    @classmethod
    def sort_team_members(cls, team_members):
        return sorted(team_members, key=lambda member: member.name)

    @field_validator('subscribers', mode="before")
    @classmethod
    def convert_subscribers(cls, v):
        if not v or not isinstance(v, list):
            return v

        converted_subscribers = []
        for subscriber in v:
            if isinstance(subscriber, ObjectId):
                converted_subscribers.append(UserWithoutTenantsDTO.from_mongo_reference_field(subscriber))
            elif isinstance(subscriber, UserWithoutTenantsDTO):
                converted_subscribers.append(subscriber)
            elif isinstance(subscriber, dict):
                converted_subscribers.append(UserWithoutTenantsDTO(**subscriber))
            else:
                converted_subscribers.append(subscriber)
        return converted_subscribers

    @field_validator('deleted_by', mode="before")
    @classmethod
    def convert_deleted_by(cls, value):
        if isinstance(value, ObjectId):
            return UserWithoutTenantsDTO.from_mongo_reference_field(value)
        if isinstance(value, User):
            return mongo_to_pydantic(value, UserWithoutTenantsDTO)
        return value


class NotificationTypeDTO(BaseModel):
    identifier: str
    label: str
    description: str


class TeamSubscriptionPreferenceDTO(BaseModel):
    user: UserWithoutTenantsDTO
    notification_types: List[str]


class NotificationSubscriptionUpdateDTO(BaseModel):
    notification_types: List[str]
    user_id: Optional[str] = None

    @field_validator("notification_types", mode="before")
    @classmethod
    def validate_notification_types(cls, value: List[str]):
        if value is None:
            return []
        try:
            return ensure_valid_notification_types(value)
        except ValueError as exc:
            raise ValueError(str(exc)) from exc


@router.get("/notification-types", response_model=List[NotificationTypeDTO])
async def list_available_notification_types(
        current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
        tenant: Annotated[Tenant, Depends(get_tenant)],
):
    return [
        NotificationTypeDTO(
            identifier=definition.identifier,
            label=definition.label,
            description=definition.description,
        )
        for definition in list_notification_types()
    ]


class DayAuditDTO(BaseModel):
    id: str = Field(alias="_id", default=None)
    timestamp: datetime.datetime
    date: datetime.date
    user: UserWithoutTenantsDTO | None = None
    member_uid: str
    old_day_types: List[DayTypeReadDTO] = []
    old_comment: str = ''
    new_day_types: List[DayTypeReadDTO] = []
    new_comment: str = ''
    action: str

    @field_serializer("timestamp", when_used="json")
    def serialize_timestamp(self, timestamp: datetime.datetime) -> str:
        """Format timestamp as UTC ISO string with trailing Z."""
        return (
            timestamp.astimezone(datetime.timezone.utc)
            .isoformat(timespec="microseconds")
            .replace("+00:00", "Z")
        )

    @field_validator('user', mode="before")
    @classmethod
    def convert_user(cls, v):
        if isinstance(v, ObjectId):
            return UserWithoutTenantsDTO.from_mongo_reference_field(v)
        return v

    @field_validator('old_day_types', 'new_day_types', mode="before")
    @classmethod
    def convert_day_types(cls, v):
        if v and isinstance(v, list):
            converted = []
            for dt in v:
                if isinstance(dt, ObjectId):
                    converted.append(DayTypeReadDTO.from_mongo_reference_field(dt))
            return converted
        return v


@lru_cache(maxsize=32)
def get_holidays(tenant, year: int = datetime.datetime.now().year) -> dict:
    countries = get_unique_countries(tenant)
    holidays_dict = {}
    for country in countries:
        country_holidays_obj = get_country_holidays(country, year)
        if country == "Sweden":
            country_holidays = {date: name for date, name in country_holidays_obj.items() if
                                name not in ["Söndag", "Sunday"]}
        else:
            country_holidays = dict(country_holidays_obj)
        holidays_dict.update({country: country_holidays})
    return holidays_dict


def team_to_read_dto(team: Team, include_archived_members: bool = False) -> TeamReadDTO:
    member_dtos = [
        mongo_to_pydantic(member, TeamMemberReadDTO)
        for member in team.members(include_archived=include_archived_members)
    ]
    team_dict = {
        "_id": str(team.id),
        "name": team.name,
        "available_day_types": [
            mongo_to_pydantic(day_type, DayTypeReadDTO)
            for day_type in team.available_day_types
        ],
        "team_members": member_dtos,
        "subscribers": [
            mongo_to_pydantic(subscriber, UserWithoutTenantsDTO)
            for subscriber in team.list_subscribers()
        ],
        "is_deleted": team.is_deleted,
        "deleted_at": team.deleted_at,
        "deleted_by": team.deleted_by,
    }
    return TeamReadDTO(**team_dict)


@router.get("")
async def list_teams(current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                     tenant: Annotated[Tenant, Depends(get_tenant)],
                     include_archived: bool = Query(False),
                     include_archived_members: bool = Query(False)):
    start_time = time.perf_counter()
    teams_qs = Team.objects_with_deleted(tenant=tenant) if include_archived else Team.objects(tenant=tenant)
    teams_list = teams_qs.order_by("name")
    converter = partial(team_to_read_dto, include_archived_members=include_archived_members)
    teams = {"teams": await asyncio.gather(
        *(run_in_threadpool(converter, team) for team in teams_list))}
    print("teams preparation " + str(time.perf_counter() - start_time))
    return teams | {"holidays": get_holidays(tenant)} | await get_all_day_types(current_user, tenant)


@router.post("/{team_id}/members")
async def add_team_member(team_id: str, team_member_dto: TeamMemberWriteDTO,
                          current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                          tenant: Annotated[Tenant, Depends(get_tenant)]):
    team_member_data = team_member_dto.model_dump()
    team_member = TeamMember(**team_member_data)
    team = Team.objects(tenant=tenant, id=team_id).first()
    team.team_members.append(team_member)
    team.save()
    tenant.update_max_team_members_in_the_period()
    return {"message": "Team member created successfully"}


@router.post("")
async def add_team(team_dto: TeamWriteDTO, current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                   tenant: Annotated[Tenant, Depends(get_tenant)]):
    team_data = team_dto.model_dump()
    team_data.update({"tenant": tenant})
    Team(**team_data).save()
    return {"message": "Team created successfully"}


@router.delete("/{team_id}")
async def delete_team(team_id: str, current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                      tenant: Annotated[Tenant, Depends(get_tenant)]):
    team = Team.objects_with_deleted(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    if not team.is_deleted:
        team.is_deleted = True
        team.deleted_at = datetime.datetime.now(datetime.timezone.utc)
        team.deleted_by = current_user
        team.save()
    return RedirectResponse(url="/", status_code=status.HTTP_303_SEE_OTHER)


@router.delete("/{team_id}/members/{team_member_id}")
async def delete_team_member(team_id: str, team_member_id: str,
                             current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                             tenant: Annotated[Tenant, Depends(get_tenant)]):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    team_member_to_remove = team.get_member(team_member_id, include_archived=True)
    if not team_member_to_remove:
        raise HTTPException(status_code=404, detail="Team member not found")
    if not getattr(team_member_to_remove, "is_deleted", False):
        team_member_to_remove.is_deleted = True
        team_member_to_remove.deleted_at = datetime.datetime.now(datetime.timezone.utc)
        team_member_to_remove.deleted_by = current_user
        team.save()
    return RedirectResponse(url="/", status_code=status.HTTP_303_SEE_OTHER)


@router.put("/{team_id}")
async def update_team(team_id: str, team_dto: TeamWriteDTO,
                      current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                      tenant: Annotated[Tenant, Depends(get_tenant)]):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if team:
        team.name = team_dto.name
        team.available_day_types = team_dto.available_day_types
        team.save()
        return {"message": "Team modified successfully"}
    else:
        raise HTTPException(status_code=404, detail="Team not found")


@router.put("/{team_id}/members/{team_member_id}")
async def update_team_member(team_id: str, team_member_id: str, team_member_dto: TeamMemberWriteDTO,
                             current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                             tenant: Annotated[Tenant, Depends(get_tenant)]):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    team_member: TeamMember | None = team.get_member(team_member_id, include_archived=True)
    if not team_member:
        raise HTTPException(status_code=404, detail="Team member not found")
    if getattr(team_member, "is_deleted", False):
        raise HTTPException(status_code=400, detail="Team member is archived")

    team_member.name = team_member_dto.name
    team_member.country = team_member_dto.country
    team_member.email = team_member_dto.email if team_member_dto.email else None
    team_member.phone = team_member_dto.phone
    team_member.birthday = team_member_dto.birthday
    team_member.available_day_types = team_member_dto.available_day_types
    team_member.employee_start_date = team_member_dto.employee_start_date
    team_member.yearly_vacation_days = team_member_dto.yearly_vacation_days

    team.save()
    return {"message": "Team member modified successfully"}


@router.post("/{team_id}/subscribe")
async def subscribe_user_to_team(
        team_id: str,
        current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
        tenant: Annotated[Tenant, Depends(get_tenant)],
        user_id: Optional[str] = None
):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    user_to_subscribe = User.objects(id=user_id).first() if user_id else current_user

    if team.is_subscribed(user_to_subscribe):
        raise HTTPException(status_code=400, detail="User already subscribed to the team")

    team.notification_preferences[str(user_to_subscribe.id)] = list_notification_type_ids()
    team.save()
    return {"message": "User subscribed to the team successfully"}


@router.post("/{team_id}/unsubscribe")
async def unsubscribe_user_from_team(
        team_id: str,
        current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
        tenant: Annotated[Tenant, Depends(get_tenant)],
        user_id: Optional[str] = None,
):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    user_to_unsubscribe = User.objects(id=user_id).first() if user_id else current_user

    subscriber_key = str(user_to_unsubscribe.id)
    if subscriber_key not in team.notification_preferences:
        raise HTTPException(status_code=400, detail="User is not subscribed to the team")

    team.notification_preferences.pop(subscriber_key, None)
    team.save()
    return {"message": "User unsubscribed from the team successfully"}


@router.get("/{team_id}/subscribers", response_model=List[UserWithoutTenantsDTO])
async def list_team_subscribers(
        team_id: str,
        current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
        tenant: Annotated[Tenant, Depends(get_tenant)]
):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    subscribers = [
        mongo_to_pydantic(subscriber, UserWithoutTenantsDTO)
        for subscriber in team.list_subscribers()
    ]

    return subscribers


@router.get("/{team_id}/notification-preferences", response_model=List[TeamSubscriptionPreferenceDTO])
async def get_team_notification_preferences(
        team_id: str,
        current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
        tenant: Annotated[Tenant, Depends(get_tenant)],
):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    default_types = list_notification_type_ids()
    preferences: List[TeamSubscriptionPreferenceDTO] = []
    for subscriber in team.list_subscribers():
        types = team.notification_preferences.get(str(subscriber.id))
        effective_types = list(types) if types is not None else list(default_types)
        preferences.append(
            TeamSubscriptionPreferenceDTO(
                user=mongo_to_pydantic(subscriber, UserWithoutTenantsDTO),
                notification_types=effective_types,
            )
        )

    return preferences


@router.put("/{team_id}/notification-preferences")
async def update_team_notification_preferences(
        team_id: str,
        subscription_update: NotificationSubscriptionUpdateDTO,
        current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
        tenant: Annotated[Tenant, Depends(get_tenant)],
):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    if subscription_update.user_id:
        user = User.objects(id=subscription_update.user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
    else:
        user = current_user

    normalized_types = subscription_update.notification_types
    subscriber_key = str(user.id)
    if normalized_types:
        team.notification_preferences[subscriber_key] = normalized_types
        team.save()
        return {
            "message": "Notification preferences updated",
            "notification_types": normalized_types,
        }

    team.notification_preferences.pop(subscriber_key, None)
    team.save()
    return {
        "message": "User unsubscribed from all notification types",
        "notification_types": [],
    }


@router.post("/move-member/{team_member_uid}")
async def transfer_team_member(current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                               tenant: Annotated[Tenant, Depends(get_tenant)],
                               team_member_uid: str,
                               target_team_id: str = Body(...),
                               source_team_id: str = Body(...)):
    source_team = Team.objects(tenant=tenant, id=source_team_id).first()
    target_team = Team.objects(tenant=tenant, id=target_team_id).first()

    if not source_team or not target_team:
        raise HTTPException(status_code=404, detail="One or both teams not found")

    team_member = source_team.get_member(team_member_uid)
    if not team_member:
        raise HTTPException(status_code=404, detail="Team member not found in source team")

    source_team.team_members = [member for member in source_team.team_members if member.uid != team_member.uid]
    source_team.save()

    if not target_team.get_member(team_member.uid, include_archived=True):
        target_team.team_members.append(team_member)
        target_team.save()

    return {"message": "Team member successfully moved"}


def is_weekend(date):
    return date.weekday() >= 5


def get_working_days(start_date, end_date, country_holidays):
    working_days = 0
    current_date = start_date
    while current_date <= end_date:
        if not is_weekend(current_date) and current_date not in country_holidays:
            working_days += 1
        current_date += datetime.timedelta(days=1)
    return working_days


def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length


@router.get("/export-absences")
async def export_absence_report(current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                                tenant: Annotated[Tenant, Depends(get_tenant)],
                                start_date: datetime.date = Query(...), end_date: datetime.date = Query(...),
                                team_ids: List[str] | None = Query(None)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Day Type Report"

    day_type_names = sorted(DayType.objects(tenant=tenant).distinct("name"))

    headers = ["Team", "Team Member Name", "Country", "Absence Days", "Working Days", "Days Worked",
               "Hours Worked"] + day_type_names
    ws.append(headers)

    body_rows = await get_report_body_rows(tenant, start_date, end_date, day_type_names, team_ids)
    for r in body_rows:
        ws.append(r)

    last_column_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_column_letter}1"
    auto_adjust_column_width(ws)

    b_io = BytesIO()
    wb.save(b_io)
    b_io.seek(0)

    return StreamingResponse(
        b_io,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=absences_{start_date}_{end_date}.xlsx"
        },
    )


def build_team_calendar(team: Team) -> Calendar:
    cal = Calendar()
    # Add calendar name so external clients display a meaningful title
    cal.extra.append(ContentLine(
        "X-WR-CALNAME",
        value=f"{team.name} - {team.tenant.name} - Vacal",
    ))
    for member in sorted(team.members(), key=lambda m: m.name):
        for date_str in sorted(member.days.keys()):
            day_entry = member.days[date_str]
            date = datetime.date.fromisoformat(date_str)
            for day_type in day_entry.day_types:
                event = Event()
                event.name = f"{member.name} - {day_type.name}"
                event.begin = date
                event.make_all_day()
                if day_entry.comment:
                    event.description = day_entry.comment
                event.uid = f"{team.id}-{member.uid}-{date_str}-{day_type.id}"
                cal.events.add(event)
    return cal


@router.get("/calendar/{team_id}")
async def get_calendar_feed(team_id: str, user_api_key: str | None = Query(None)):
    if not team_id or not user_api_key:
        raise HTTPException(status_code=404, detail="Calendar not found")
    team = Team.objects(id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Calendar not found")
    user = User.objects(auth_details__api_key=user_api_key).first()
    if not user or user.disabled:
        raise HTTPException(status_code=404, detail="Calendar not found")
    cal = build_team_calendar(team)
    return Response(cal.serialize(), media_type="text/calendar")


async def get_report_body_rows(tenant, start_date, end_date, day_type_names, team_ids: List[str] | None = None):
    body_rows = []
    country_holidays = get_holidays(tenant)
    working_hours_in_a_day = 8
    teams_qs = Team.objects(tenant=tenant).order_by("name")
    if team_ids:
        teams_qs = teams_qs.filter(id__in=team_ids)
    for team in teams_qs:
        for member in team.members():
            day_type_counts = {}
            member_holidays = country_holidays.get(member.country, [])
            absence_days_count = 0
            for date_str, day_entry in member.days.items():
                date = datetime.date.fromisoformat(date_str)
                if start_date <= date <= end_date:
                    is_absence_day = False
                    for day_type in day_entry.day_types:
                        if day_type.is_absence:
                            is_absence_day = True
                        day_type_name = day_type.name
                        if day_type_name in day_type_counts:
                            day_type_counts[day_type_name] += 1
                        else:
                            day_type_counts[day_type_name] = 1
                    if is_absence_day:
                        absence_days_count += 1
            working_days = get_working_days(start_date, end_date, member_holidays)
            body_rows.append(
                [team.name, member.name, member.country, absence_days_count, working_days,
                 working_days - absence_days_count,
                 (working_days - absence_days_count) * working_hours_in_a_day] +
                [day_type_counts.get(n, 0) for n in day_type_names])
    return body_rows


def validate_date(date_str):
    try:
        datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Can't parse date {date_str}")


def filter_out_birthdays(tenant: Tenant, day_type_ids: List[str]) -> Generator:
    birthday_day_type_id: str = DayType.get_birthday_day_type_id(tenant)
    day_type_ids = set(day_type_ids)
    return filter(lambda x: x != birthday_day_type_id, day_type_ids)


@router.put("/{team_id}/members/{team_member_id}/days")
async def update_days(team_id: str, team_member_id: str, days: Dict[str, Dict[str, str | List[str]]],
                      current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                      tenant: Annotated[Tenant, Depends(get_tenant)]):
    team: Team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    team_member: TeamMember | None = team.get_member(team_member_id, include_archived=True)
    if not team_member:
        raise HTTPException(status_code=404, detail="Team member not found")
    if getattr(team_member, "is_deleted", False):
        raise HTTPException(status_code=400, detail="Team member is archived")

    updated_days = {}
    enforce_absence_limit = current_user.role == "employee"
    audits: List[DayAudit] = []
    for date_str, day_entry_dto in days.items():
        validate_date(date_str)
        filtered_ids = list(filter_out_birthdays(tenant, day_entry_dto["day_types"]))
        day_types = list(DayType.objects(tenant=tenant, id__in=filtered_ids).order_by("name"))
        old_entry: DayEntry | None = team_member.days.get(date_str)
        new_comment = day_entry_dto.get("comment", '')

        if not day_types and new_comment == "":
            if old_entry:
                audits.append(DayAudit(
                    tenant=tenant,
                    team=team,
                    member_uid=str(team_member.uid),
                    date=datetime.date.fromisoformat(date_str),
                    user=current_user,
                    old_day_types=list(old_entry.day_types),
                    old_comment=old_entry.comment or "",
                    new_day_types=[],
                    new_comment="",
                    action="deleted",
                ))
                team_member.days.pop(date_str, None)
            continue

        day_entry = DayEntry()
        day_entry.day_types = day_types
        day_entry.comment = new_comment
        updated_days[date_str] = day_entry

        if (
            enforce_absence_limit
            and len(team.members()) > 1
            and any(day_type.is_absence for day_type in day_types)
        ):
            all_members_absent = True
            for member in team.members():
                if member.uid == team_member.uid:
                    entry = day_entry
                else:
                    entry = member.days.get(date_str)

                if not entry or not entry.day_types:
                    all_members_absent = False
                    break

                if not any(getattr(day_type, "is_absence", False) for day_type in entry.day_types):
                    all_members_absent = False
                    break

            if all_members_absent:
                raise HTTPException(
                    status_code=400,
                    detail="At least one teammate must remain available for this day.",
                )

        audits.append(DayAudit(
            tenant=tenant,
            team=team,
            member_uid=str(team_member.uid),
            date=datetime.date.fromisoformat(date_str),
            user=current_user,
            old_day_types=list(old_entry.day_types) if old_entry else [],
            old_comment=old_entry.comment if old_entry else "",
            new_day_types=list(day_entry.day_types),
            new_comment=day_entry.comment,
            action="created" if old_entry is None else "updated",
        ))

    if updated_days:
        if not team_member.days:
            team_member.days = updated_days
        else:
            team_member.days.update(updated_days)

    team.save()
    for audit in audits:
        audit.save()

    return {"message": "Days modified successfully"}


@router.get("/{team_id}/members/{team_member_id}/days/{date}/history")
async def get_day_history(
    team_id: str,
    team_member_id: str,
    date: str,
    current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
    tenant: Annotated[Tenant, Depends(get_tenant)],
    skip: int = 0,
    limit: int = 100,
):
    team: Team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    team_member: TeamMember | None = team.get_member(team_member_id, include_archived=True)
    if not team_member:
        raise HTTPException(status_code=404, detail="Team member not found")
    validate_date(date)
    date_obj = datetime.datetime.strptime(date, "%Y-%m-%d").date()
    audits = _get_paginated_audits(tenant, team, team_member, skip, limit, date_obj)
    return [mongo_to_pydantic(a, DayAuditDTO) for a in audits]


@router.get("/{team_id}/members/{team_member_id}/history")
async def get_member_history(
    team_id: str,
    team_member_id: str,
    current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
    tenant: Annotated[Tenant, Depends(get_tenant)],
    skip: int = 0,
    limit: int = 100,
):
    team: Team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    team_member: TeamMember | None = team.get_member(team_member_id, include_archived=True)
    if not team_member:
        raise HTTPException(status_code=404, detail="Team member not found")
    audits = _get_paginated_audits(tenant, team, team_member, skip, limit)
    return [mongo_to_pydantic(a, DayAuditDTO) for a in audits]


def _get_paginated_audits(
    tenant: Tenant,
    team: Team,
    team_member: TeamMember,
    skip: int,
    limit: int,
    date: datetime.date | None = None,
):
    skip = max(skip, 0)
    limit = min(limit, 100)
    query = DayAudit.objects(
        tenant=tenant, team=team, member_uid=str(team_member.uid)
    )
    if date is not None:
        query = query.filter(date=date)
    return query.order_by("-timestamp").skip(skip).limit(limit)
