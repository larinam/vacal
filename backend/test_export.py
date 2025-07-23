import os
os.environ.setdefault("MONGO_MOCK", "1")
os.environ.setdefault("AUTHENTICATION_SECRET_KEY", "test_secret")

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from .main import app
from .model import Tenant, DayType, Team, TeamMember, DayEntry, User
import uuid

client = TestClient(app)


def setup_teams():
    tenant = Tenant(name=f"Tenant{uuid.uuid4()}", identifier=str(uuid.uuid4())).save()
    DayType.init_day_types(tenant)
    vacation = DayType.objects(tenant=tenant, identifier="vacation").first()
    member1 = TeamMember(name="Alice", country="Sweden", days={"2025-01-01": DayEntry(day_types=[vacation])})
    team1 = Team(tenant=tenant, name="Team1", team_members=[member1])
    team1.save()
    member2 = TeamMember(name="Bob", country="Sweden", days={"2025-01-01": DayEntry(day_types=[vacation])})
    team2 = Team(tenant=tenant, name="Team2", team_members=[member2])
    team2.save()
    return team1, team2


from backend.dependencies import get_current_active_user_check_tenant, get_tenant


def test_export_selected_team():
    team1, team2 = setup_teams()
    app.dependency_overrides[get_current_active_user_check_tenant] = lambda: User(tenants=[team1.tenant])
    app.dependency_overrides[get_tenant] = lambda: team1.tenant
    resp = client.get(f"/teams/export-vacations?start_date=2025-01-01&end_date=2025-12-31&team_ids={team1.id}")
    assert resp.status_code == 200
    app.dependency_overrides = {}
    from io import BytesIO
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert any(r[0] == "Team1" for r in rows[1:])
    assert not any(r[0] == "Team2" for r in rows[1:])
