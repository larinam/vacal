import datetime
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.dependencies import get_current_active_user_check_tenant, get_tenant
from backend.main import app
from backend.model import DayEntry, DayType, Team, TeamMember, Tenant, User
from backend.routers.teams import get_holidays, get_working_days

client = TestClient(app)


def setup_teams():
    tenant = Tenant(name=f"Tenant{uuid.uuid4()}", identifier=str(uuid.uuid4())).save()
    DayType.init_day_types(tenant)
    vacation = DayType.objects(tenant=tenant, identifier="vacation").first()
    comp_leave = DayType.objects(tenant=tenant, identifier="compensatory_leave").first()
    member1 = TeamMember(
        name="Alice",
        country="Sweden",
        days={
            "2025-01-01": DayEntry(day_types=[vacation]),
            "2025-01-02": DayEntry(day_types=[comp_leave]),
        },
    )
    team1 = Team(tenant=tenant, name="Team1", team_members=[member1])
    team1.save()
    member2 = TeamMember(name="Bob", country="Sweden", days={"2025-01-01": DayEntry(day_types=[vacation])})
    team2 = Team(tenant=tenant, name="Team2", team_members=[member2])
    team2.save()
    return team1, team2


def test_export_selected_team():
    team1, team2 = setup_teams()
    app.dependency_overrides[get_current_active_user_check_tenant] = lambda: User(tenants=[team1.tenant])
    app.dependency_overrides[get_tenant] = lambda: team1.tenant
    resp = client.get(
        f"/teams/export-absences?start_date=2025-01-01&end_date=2025-12-31&team_ids={team1.id}"
    )
    assert resp.status_code == 200
    assert (
        resp.headers["Content-Disposition"]
        == "attachment; filename=absences_2025-01-01_2025-12-31.xlsx"
    )
    app.dependency_overrides = {}
    from io import BytesIO
    wb = load_workbook(BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    assert "Absence Days" in headers
    assert "Vacation" in headers
    assert "Compensatory leave" in headers
    abs_idx = headers.index("Absence Days")
    vac_idx = headers.index("Vacation")
    comp_idx = headers.index("Compensatory leave")
    assert any(r[0] == "Team1" for r in rows[1:])
    assert not any(r[0] == "Team2" for r in rows[1:])
    team1_row = next(r for r in rows[1:] if r[0] == "Team1")
    assert team1_row[abs_idx] == 2
    assert team1_row[vac_idx] == 1
    assert team1_row[comp_idx] == 1


def test_export_includes_deleted_member_until_last_working_day():
    tenant = Tenant(name=f"Tenant{uuid.uuid4()}", identifier=str(uuid.uuid4())).save()
    DayType.init_day_types(tenant)
    vacation = DayType.objects(tenant=tenant, identifier="vacation").first()

    deleted_member = TeamMember(
        name="Charlie",
        country="Sweden",
        days={
            "2025-01-10": DayEntry(day_types=[vacation]),
            "2025-01-20": DayEntry(day_types=[vacation]),
        },
        last_working_day=datetime.date(2025, 1, 15),
        is_deleted=True,
    )

    team = Team(tenant=tenant, name="TeamDeleted", team_members=[deleted_member])
    team.save()

    app.dependency_overrides[get_current_active_user_check_tenant] = lambda: User(tenants=[tenant])
    app.dependency_overrides[get_tenant] = lambda: tenant

    response = client.get(
        f"/teams/export-absences?start_date=2025-01-01&end_date=2025-01-31&team_ids={team.id}"
    )

    app.dependency_overrides = {}

    assert response.status_code == 200

    from io import BytesIO

    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    headers = rows[0]
    working_idx = headers.index("Working Days")
    absence_idx = headers.index("Absence Days")
    vacation_idx = headers.index("Vacation")

    member_row = next(r for r in rows[1:] if r[1] == "Charlie")

    holidays = get_holidays(tenant)
    effective_end = datetime.date(2025, 1, 15)
    expected_working_days = get_working_days(
        datetime.date(2025, 1, 1), effective_end, holidays.get("Sweden", {})
    )

    assert member_row[working_idx] == expected_working_days
    assert member_row[absence_idx] == 1
    assert member_row[vacation_idx] == 1


def test_export_excludes_deleted_member_before_period():
    tenant = Tenant(name=f"Tenant{uuid.uuid4()}", identifier=str(uuid.uuid4())).save()
    DayType.init_day_types(tenant)
    vacation = DayType.objects(tenant=tenant, identifier="vacation").first()

    deleted_member = TeamMember(
        name="Dana",
        country="Sweden",
        days={"2024-12-15": DayEntry(day_types=[vacation])},
        last_working_day=datetime.date(2024, 12, 31),
        is_deleted=True,
    )

    team = Team(tenant=tenant, name="TeamArchive", team_members=[deleted_member])
    team.save()

    app.dependency_overrides[get_current_active_user_check_tenant] = lambda: User(tenants=[tenant])
    app.dependency_overrides[get_tenant] = lambda: tenant

    response = client.get(
        f"/teams/export-absences?start_date=2025-01-01&end_date=2025-01-31&team_ids={team.id}"
    )

    app.dependency_overrides = {}

    assert response.status_code == 200

    from io import BytesIO

    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    assert all(r[1] != "Dana" for r in rows[1:])
