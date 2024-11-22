import asyncio
import datetime
import hashlib
import hmac
import logging
import os
import time
from collections import defaultdict
from contextlib import asynccontextmanager
from copy import deepcopy
from io import BytesIO
from typing import List, Dict, Annotated, Self, Generator, Optional

import pycountry
from apscheduler.schedulers.background import BackgroundScheduler
from bson import ObjectId
from fastapi import FastAPI, status, Body, Depends
from fastapi import Query, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from prometheus_fastapi_instrumentator import Instrumentator
from pycountry.db import Country
from pydantic import BaseModel, Field, computed_field, EmailStr
from pydantic.functional_validators import field_validator, model_validator
from starlette.concurrency import run_in_threadpool

from .dependencies import create_access_token, get_current_active_user_check_tenant, get_tenant, mongo_to_pydantic, \
    TenantMiddleware, tenant_var
from .model import Team, TeamMember, get_unique_countries, DayType, User, Tenant, DayEntry
from .routers import users, daytypes, management
from .routers.daytypes import DayTypeReadDTO, get_all_day_types
from .routers.users import UserWithoutTenantsDTO
from .sheduled.activate_trials import activate_trials
from .sheduled.birthdays import send_birthday_email_updates
from .sheduled.update_max_team_members_numbers import run_update_max_team_members_numbers
from .sheduled.vacation_starts import send_vacation_email_updates
from .utils import get_country_holidays

origins = [
    "http://localhost",
    "http://localhost:3000",
    "http://127.0.0.1",
    "http://127.0.0.1:3000",
]

cors_origin = os.getenv("CORS_ORIGIN")  # should contain production domain of the frontend
if cors_origin:  # for production
    origins.append(cors_origin)

ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 365

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_BOT_USERNAME = os.getenv("TELEGRAM_BOT_USERNAME", "")

MULTITENANCY_ENABLED = os.getenv("MULTITENANCY_ENABLED", False)


@asynccontextmanager
async def lifespan(app: FastAPI):
    scheduler = BackgroundScheduler()
    scheduler.add_job(send_vacation_email_updates, 'cron', hour=6, minute=0)
    scheduler.add_job(send_birthday_email_updates, 'cron', hour=6, minute=5)
    if MULTITENANCY_ENABLED:
        scheduler.add_job(run_update_max_team_members_numbers, 'cron', hour=1, minute=5)
        scheduler.add_job(activate_trials, 'cron', hour=2, minute=5)
    scheduler.start()
    yield
    scheduler.shutdown()


app = FastAPI(lifespan=lifespan)
app.add_middleware(TenantMiddleware)
app.include_router(users.router)
app.include_router(daytypes.router)
if MULTITENANCY_ENABLED:
    app.include_router(management.router)
Instrumentator().instrument(app).expose(app)

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

log = logging.getLogger(__name__)


class GeneralApplicationConfigDTO(BaseModel):
    telegram_enabled: bool
    telegram_bot_username: str
    user_initiated: bool
    multitenancy_enabled: bool = False


# noinspection PyNestedDecorators
class TeamMemberWriteDTO(BaseModel):
    name: str
    country: str
    email: EmailStr | None = None
    phone: str | None = None
    available_day_types: List[DayTypeReadDTO] = []
    birthday: str | None = Field(None, pattern=r"^(0[1-9]|1[0-2])-(0[1-9]|[12][0-9]|3[01])$")

    @field_validator("country")
    @classmethod
    def validate_country(cls, value: str) -> str:
        country_name = validate_country_name(value)
        if country_name:
            return country_name
        raise ValueError("Invalid country name")

    @field_validator('email', mode='before')
    @classmethod
    def empty_email_to_none(cls, v):
        return None if v == "" else v

    @field_validator('birthday', mode='before')
    @classmethod
    def empty_birthday_to_none(cls, v):
        return None if v == "" else v


class DayEntryDTO(BaseModel):
    day_types: List[DayTypeReadDTO] = []
    comment: str = ''


# noinspection PyNestedDecorators
class TeamMemberReadDTO(TeamMemberWriteDTO):
    uid: str
    days: Dict[str, DayEntryDTO] = Field(default_factory=dict)

    @computed_field
    @property
    def vacation_days_by_year(self) -> Dict[int, int]:
        vac_days_count = defaultdict(int)
        vacation_day_type = mongo_to_pydantic(DayType.objects(tenant=tenant_var.get(), name="Vacation").first(),
                                              DayTypeReadDTO)
        for date_str, day_entry in self.days.items():
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            day_types = day_entry.day_types
            if vacation_day_type in day_types:
                vac_days_count[date.year] += 1
        return dict(vac_days_count)

    @model_validator(mode='after')
    def include_birthday(self) -> Self:
        def add_birthday(year):
            birthday_date = f"{year}-{self.birthday}"
            if birthday_date not in self.days:
                self.days[birthday_date] = DayEntryDTO()
            self.days[birthday_date].day_types.append(
                DayTypeReadDTO.from_mongo_reference_field(DayType.get_birthday_day_type_id(tenant_var.get())))

        if self.birthday:
            current_year = datetime.datetime.now().year
            add_birthday(current_year)  # Add for current year
            add_birthday(current_year + 1)  # Add for next year

        return self

    @field_validator('days', mode="before")
    @classmethod
    def convert_days(cls, v):
        if v and isinstance(v, dict):
            converted_days = v
            for date_str, day_entry in v.items():
                day_type_ids = day_entry["day_types"]
                converted_day_types = []
                for day_type_id in day_type_ids:
                    if isinstance(day_type_id, ObjectId):
                        converted_day_types.append(DayTypeReadDTO.from_mongo_reference_field(day_type_id))
                converted_days[date_str]["day_types"] = converted_day_types
            return converted_days
        return v

    @computed_field
    @property
    def country_flag(self) -> str:
        country = pycountry.countries.get(name=self.country)
        if country:
            return country.flag
        raise ValueError(f"Invalid country name: {self.country}")


class TeamWriteDTO(BaseModel):
    name: str
    available_day_types: List[DayTypeReadDTO] = []


# noinspection PyNestedDecorators
class TeamReadDTO(TeamWriteDTO):
    id: str = Field(None, alias='_id')
    team_members: List[TeamMemberReadDTO]
    subscribers: List[UserWithoutTenantsDTO] = []

    @field_validator('team_members')
    @classmethod
    def sort_team_members(cls, team_members):
        return sorted(team_members, key=lambda member: member.name)

    @field_validator('subscribers', mode="before")
    @classmethod
    def convert_subscribers(cls, v):
        if v and isinstance(v, list):
            converted_subscribers = []
            for subscriber in v:
                if isinstance(subscriber, ObjectId):
                    converted_subscribers.append(UserWithoutTenantsDTO.from_mongo_reference_field(subscriber))
            return converted_subscribers
        return v


class TokenDTO(BaseModel):
    access_token: str
    token_type: str


class TokenDataDTO(BaseModel):
    username: str | None = None


def validate_country_name(country_name):
    countries = pycountry.countries.search_fuzzy(country_name)
    if countries:
        country = countries[0]
        if isinstance(country, Country):
            return country.name
    return None


def get_holidays(tenant, year: int = datetime.datetime.now().year) -> dict:
    countries = get_unique_countries(tenant)
    holidays_dict = {}
    for country in countries:
        country_holidays_obj = get_country_holidays(country, year)
        # Filtering out holidays with "Söndag" for Sweden
        if country == "Sweden":
            country_holidays = {date: name for date, name in country_holidays_obj.items() if
                                name not in ["Söndag", "Sunday"]}
        else:
            country_holidays = dict(country_holidays_obj)
        holidays_dict.update({country: country_holidays})
    return holidays_dict


# General Application Configuration
@app.get("/config", response_model=GeneralApplicationConfigDTO)
async def get_config():
    tenant_exists = Tenant.objects().first() is not None
    user_exists = User.objects().first() is not None
    return {"telegram_enabled": bool(TELEGRAM_BOT_TOKEN and TELEGRAM_BOT_USERNAME),
            "telegram_bot_username": TELEGRAM_BOT_USERNAME,
            "user_initiated": tenant_exists and user_exists,
            "multitenancy_enabled": MULTITENANCY_ENABLED}


# Authentication
@app.post("/token")
async def login_for_access_token(form_data: Annotated[OAuth2PasswordRequestForm, Depends()]) -> TokenDTO:
    user = User.authenticate_user(form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = datetime.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.auth_details.username}, expires_delta=access_token_expires
    )
    return TokenDTO(access_token=access_token, token_type="bearer")


# noinspection PyNestedDecorators
class TelegramAuthData(BaseModel):
    hash: str
    id: int
    auth_date: int
    username: str

    # Additional fields should be added if necessary for further processing in the endpoint method

    @field_validator('auth_date')
    @classmethod
    def check_auth_date(cls, auth_date):
        if (time.time() - auth_date) > 86400:
            raise ValueError("Data is outdated")
        return auth_date

    @model_validator(mode="before")
    @classmethod
    def check_telegram_authorization(cls, input_values):
        values = deepcopy(input_values)
        telegram_hash = values.pop("hash")
        data_check_arr = ["{}={}".format(key, value) for key, value in values.items()]
        data_check_arr.sort()
        data_check_string = "\n".join(data_check_arr)

        secret_key = hashlib.sha256(TELEGRAM_BOT_TOKEN.encode()).digest()
        calculated_hash = hmac.new(secret_key, data_check_string.encode(), hashlib.sha256).hexdigest()

        if calculated_hash != telegram_hash:
            raise ValueError("Data is NOT from Telegram")

        return input_values


@app.post("/telegram-login")
async def telegram_login(auth_data: TelegramAuthData):
    # At this point, auth_data is already validated by Pydantic
    username = auth_data.dict().get("username").lower()  # Telegram sends the username in the auth data
    user = User.get_by_telegram_username(username)

    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"The user with your Telegram username: {username} is not found in our system"
        )
    # update Telegram ID
    if not user.auth_details.telegram_id:
        user.auth_details.telegram_id = auth_data.dict().get("id")
        user.save()
    access_token_expires = datetime.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.auth_details.username}, expires_delta=access_token_expires
    )

    return TokenDTO(access_token=access_token, token_type="bearer")


# Business Logic
@app.get("/")
async def read_root(current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                    tenant: Annotated[Tenant, Depends(get_tenant)]):
    start_time = time.perf_counter()
    teams_list = Team.objects(tenant=tenant).order_by("name")
    teams = {"teams": await asyncio.gather(
        *(run_in_threadpool(mongo_to_pydantic, team, TeamReadDTO) for team in teams_list))}
    print("teams preparation " + str(time.perf_counter() - start_time))
    return teams | {"holidays": get_holidays(tenant)} | await get_all_day_types(current_user, tenant)


@app.post("/teams/{team_id}/members")
async def add_team_member(team_id: str, team_member_dto: TeamMemberWriteDTO,
                          current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                          tenant: Annotated[Tenant, Depends(get_tenant)]):
    team_member_data = team_member_dto.model_dump()
    team_member = TeamMember(**team_member_data)
    team = Team.objects(tenant=tenant, id=team_id).first()
    team.team_members.append(team_member)
    team.save()
    tenant.update_max_team_members_in_the_period()
    return {"message": "Team member created successfully"}


@app.post("/teams")
async def add_team(team_dto: TeamWriteDTO, current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                   tenant: Annotated[Tenant, Depends(get_tenant)]):
    team_data = team_dto.model_dump()
    team_data.update({"tenant": tenant})
    Team(**team_data).save()
    return {"message": "Team created successfully"}


@app.delete("/teams/{team_id}")
async def delete_team(team_id: str, current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                      tenant: Annotated[Tenant, Depends(get_tenant)]):
    Team.objects(tenant=tenant, id=team_id).delete()
    return RedirectResponse(url="/", status_code=status.HTTP_303_SEE_OTHER)


@app.delete("/teams/{team_id}/members/{team_member_id}")
async def delete_team_member(team_id: str, team_member_id: str,
                             current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                             tenant: Annotated[Tenant, Depends(get_tenant)]):
    team = Team.objects(tenant=tenant, id=team_id).first()
    team_members = team.team_members
    team_member_to_remove = team_members.get(uid=team_member_id)
    team_members.remove(team_member_to_remove)
    team.save()
    return RedirectResponse(url="/", status_code=status.HTTP_303_SEE_OTHER)


@app.put("/teams/{team_id}")
async def update_team(team_id: str, team_dto: TeamWriteDTO,
                      current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                      tenant: Annotated[Tenant, Depends(get_tenant)]):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if team:
        team.name = team_dto.name
        team.available_day_types = team_dto.available_day_types
        team.save()
        return {"message": "Team modified successfully"}
    else:
        raise HTTPException(status_code=404, detail="Team not found")


@app.put("/teams/{team_id}/members/{team_member_id}")
async def update_team_member(team_id: str, team_member_id: str, team_member_dto: TeamMemberWriteDTO,
                             current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                             tenant: Annotated[Tenant, Depends(get_tenant)]):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    team_member: TeamMember = team.team_members.get(uid=team_member_id)
    if not team_member:
        raise HTTPException(status_code=404, detail="Team member not found")

    team_member.name = team_member_dto.name
    team_member.country = team_member_dto.country
    team_member.email = team_member_dto.email if team_member_dto.email else None
    team_member.phone = team_member_dto.phone
    team_member.birthday = team_member_dto.birthday
    team_member.available_day_types = team_member_dto.available_day_types

    team.save()
    return {"message": "Team member modified successfully"}


@app.post("/teams/{team_id}/subscribe")
async def subscribe_user_to_team(
        team_id: str,
        current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
        tenant: Annotated[Tenant, Depends(get_tenant)],
        user_id: Optional[str] = None
):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    user_to_subscribe = User.objects(id=user_id).first() if user_id else current_user

    if user_to_subscribe in team.subscribers:
        raise HTTPException(status_code=400, detail="User already subscribed to the team")

    team.subscribers.append(user_to_subscribe)
    team.save()
    return {"message": "User subscribed to the team successfully"}


@app.post("/teams/{team_id}/unsubscribe")
async def unsubscribe_user_from_team(
        team_id: str,
        current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
        tenant: Annotated[Tenant, Depends(get_tenant)],
        user_id: Optional[str] = None,
):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    user_to_unsubscribe = User.objects(id=user_id).first() if user_id else current_user

    if user_to_unsubscribe not in team.subscribers:
        raise HTTPException(status_code=400, detail="User is not subscribed to the team")

    team.subscribers.remove(user_to_unsubscribe)
    team.save()
    return {"message": "User unsubscribed from the team successfully"}


@app.get("/teams/{team_id}/subscribers", response_model=List[UserWithoutTenantsDTO])
async def list_team_subscribers(
        team_id: str,
        current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
        tenant: Annotated[Tenant, Depends(get_tenant)]
):
    team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    subscribers = [
        mongo_to_pydantic(subscriber, UserWithoutTenantsDTO)
        for subscriber in team.subscribers
    ]

    return subscribers


@app.post("/move-team-member/{team_member_uid}")
async def move_team_member(current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                           tenant: Annotated[Tenant, Depends(get_tenant)],
                           team_member_uid: str,
                           target_team_id: str = Body(...),
                           source_team_id: str = Body(...)):
    # Retrieve source and target teams
    source_team = Team.objects(tenant=tenant, id=source_team_id).first()
    target_team = Team.objects(tenant=tenant, id=target_team_id).first()

    if not source_team or not target_team:
        raise HTTPException(status_code=404, detail="One or both teams not found")

    # Find the team member in source team
    team_member = next((member for member in source_team.team_members if str(member.uid) == str(team_member_uid)), None)
    if not team_member:
        raise HTTPException(status_code=404, detail="Team member not found in source team")

    # Remove team member from source team
    source_team.team_members = [member for member in source_team.team_members if member.uid != team_member.uid]
    source_team.save()

    # Add team member to target team, ensuring no duplicates
    if not any(member.uid == team_member.uid for member in target_team.team_members):
        target_team.team_members.append(team_member)
        target_team.save()

    return {"message": "Team member successfully moved"}


def is_weekend(date):
    return date.weekday() >= 5  # 5 for Saturday and 6 for Sunday


def get_working_days(start_date, end_date, country_holidays):
    # Calculate the number of working days excluding weekends and holidays
    working_days = 0
    current_date = start_date
    while current_date <= end_date:
        if not is_weekend(current_date) and current_date not in country_holidays:
            working_days += 1
        current_date += datetime.timedelta(days=1)
    return working_days


def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length


@app.get("/export-vacations")
async def export_vacations(current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                           tenant: Annotated[Tenant, Depends(get_tenant)],
                           start_date: datetime.date = Query(...), end_date: datetime.date = Query(...)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Day Type Report"

    day_type_names = sorted(DayType.objects(tenant=tenant, identifier__ne="vacation").distinct("name"))

    headers = ["Team", "Team Member Name", "Country", "Vacation Days", "Working Days", "Days Worked",
               "Hours Worked"] + day_type_names
    ws.append(headers)

    body_rows = await get_report_body_rows(tenant, start_date, end_date, day_type_names)
    for r in body_rows:
        ws.append(r)

    # Assuming you have added data below headers, now apply auto_filter
    last_column_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_column_letter}1"
    auto_adjust_column_width(ws)

    # Save the workbook to a BytesIO object
    b_io = BytesIO()
    wb.save(b_io)
    b_io.seek(0)  # Go to the start of the BytesIO object

    # Create and return the streaming response
    return StreamingResponse(b_io, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={
                                 "Content-Disposition": f"attachment; filename=vacations_{start_date}_{end_date}.xlsx"})


async def get_report_body_rows(tenant, start_date, end_date, day_type_names):
    body_rows = []
    vacation_day_type_id = DayType.get_vacation_day_type_id(tenant)
    country_holidays = get_holidays(tenant)
    working_hours_in_a_day = 8
    # Query and process the data
    for team in Team.objects(tenant=tenant):
        for member in team.team_members:
            day_type_counts = {}
            member_holidays = country_holidays.get(member.country, [])
            vac_days_count = 0
            for date_str, day_entry in member.days.items():
                # Convert the string to a date object to compare with the given date range
                date = datetime.date.fromisoformat(date_str)
                if start_date <= date <= end_date:
                    for day_type in day_entry.day_types:
                        if vacation_day_type_id == str(day_type.id):
                            # Count only if the day type list contains the 'Vacation' day type ID
                            vac_days_count += 1
                            continue

                        # Get the name of the day type
                        day_type_name = day_type.name

                        # Increment the count for this day type
                        if day_type_name in day_type_counts:
                            day_type_counts[day_type_name] += 1
                        else:
                            day_type_counts[day_type_name] = 1
            # Calculate working days
            working_days = get_working_days(start_date, end_date, member_holidays)
            body_rows.append(
                [team.name, member.name, member.country, vac_days_count, working_days, working_days - vac_days_count,
                 (working_days - vac_days_count) * working_hours_in_a_day] + [day_type_counts.get(n, 0) for n in
                                                                              day_type_names])
    return body_rows


def validate_date(date_str):
    try:
        datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Can't parse date {date_str}")


def filter_out_birthdays(tenant: Tenant, day_type_ids: List[str]) -> Generator:
    birthday_day_type_id: str = DayType.get_birthday_day_type_id(tenant)
    day_type_ids = set(day_type_ids)
    return filter(lambda x: x != birthday_day_type_id, day_type_ids)


@app.put("/teams/{team_id}/members/{team_member_id}/days")
async def update_days(team_id: str, team_member_id: str, days: Dict[str, Dict[str, str | List[str]]],
                      current_user: Annotated[User, Depends(get_current_active_user_check_tenant)],
                      tenant: Annotated[Tenant, Depends(get_tenant)]):
    """Assume it is an update for only one day"""
    team: Team = Team.objects(tenant=tenant, id=team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")

    team_member: TeamMember = team.team_members.get(uid=team_member_id)
    if not team_member:
        raise HTTPException(status_code=404, detail="Team member not found")

    updated_days = {}
    for date_str, day_entry_dto in days.items():
        validate_date(date_str)
        filtered_ids = filter_out_birthdays(tenant, day_entry_dto["day_types"])
        day_types = DayType.objects(tenant=tenant, id__in=filtered_ids).order_by("name")
        day_entry: DayEntry = team_member.days.get(date_str, DayEntry())
        day_entry.day_types = day_types
        day_entry.comment = day_entry_dto.get("comment", '')
        updated_days[date_str] = day_entry

    if not team_member.days:
        # otherwise breaks with mongoengine.errors.OperationError and can't add days for newly created team members
        team_member.days = updated_days
    else:
        team_member.days.update(updated_days)
    team.save()
    return {"message": "Days modified successfully"}
